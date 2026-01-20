from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO
from app_modules.cell_mapping import CELL_MAP

HEADLINE_COLORS = ["FF0BD7B5", "0BD7B5"]


def fill_excel(template_bytes, field_values, summary_text):
    wb = load_workbook(filename=BytesIO(template_bytes))

    for sheet_name, fields in CELL_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        for field_key, cell_ref in fields.items():
            value = field_values.get(field_key, "")

            cell = ws[cell_ref]

            fill = cell.fill
            if (
                fill and isinstance(fill, PatternFill)
                and fill.fgColor and fill.fgColor.rgb
                and fill.fgColor.rgb.upper() in HEADLINE_COLORS
            ):
                continue

            cell.value = value

    first_sheet = wb.sheetnames[0]
    ws_first = wb[first_sheet]

    if summary_text:
        placed = False
        for row in ws_first.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "skriv her" in cell.value.lower():
                    cell.value = summary_text
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    placed = True
                    break
            if placed:
                break

        if not placed:
            ws_first["A46"] = summary_text
            ws_first["A46"].alignment = Alignment(wrap_text=True, vertical="top")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
