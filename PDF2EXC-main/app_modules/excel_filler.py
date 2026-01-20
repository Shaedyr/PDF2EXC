from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

HEADLINE_COLORS = ["FF0BD7B5", "0BD7B5"]

def fill_excel(template_bytes, field_values, cell_map):
    """
    Generic filler: writes field_values into cells defined in cell_map.
    Does NOT handle summary text anymore.
    """
    wb = load_workbook(filename=BytesIO(template_bytes))

    for sheet_name, fields in cell_map.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        for field_key, cell_ref in fields.items():
            value = field_values.get(field_key, "")
            cell = ws[cell_ref]

            fill = cell.fill
            if (
                fill and isinstance(fill, PatternFill)
                and fill.fgColor and fill.fgColor.rgb
                and fill.fgColor.rgb.upper() in HEADLINE_COLORS
            ):
                continue

            cell.value = value

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
